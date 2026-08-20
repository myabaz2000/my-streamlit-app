import io
import os
import folium
from folium.plugins import MarkerCluster
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from streamlit_folium import st_folium

st.set_page_config(
    page_title="منظومة مكاتب التصويت - قيادة أملن",
    page_icon="🗳️",
    layout="wide"
)

st.title("🗳️ منظومة إدارة مكاتب التصويت - قيادة أملن")
st.subheader("الجماعات: أملن - تارسواط - تاسريرت (رؤساء المكاتب ونوابهم والأعضاء ونوابهم)")

# =============================================================
# 2. مسارات الملفات (مسارات نسبية تعمل على السيرفر السحابي)
# =============================================================
EXCEL_FILE = 'Bureaux_de_Votes_Ammelne_2026_Modf_2026.xlsx'
LOGISTICS_FILE = 'logistics_data.xlsx'
TRANSPORT_FILE = 'transport_data.xlsx'

# -------------------------------------------------------------
# بيانات الأعوان المكلفين المباشرة (جماعة أملن وجماعة تاسريرت)
# -------------------------------------------------------------
COMMUNES_AGENTS = {
    "أملن": {
        "1": {"agent": "بلهنا", "phone": "0623491029"},
        "2": {"agent": "مسولي", "phone": "0623491031"},
        "3": {"agent": "مسولي", "phone": "0623491031"},
        "4": {"agent": "العليم", "phone": "0623491033"},
        "5": {"agent": "بوبرا", "phone": "0611216500"},
        "6": {"agent": "حفيظ", "phone": "0662399366"},
        "7": {"agent": "حفيظ", "phone": "0623491031"},
        "8": {"agent": "حفيظ", "phone": "0623491031"},
        "9": {"agent": "الشاذلي", "phone": "0623490798"},
        "10": {"agent": "الشاذلي", "phone": "0623490798"},
        "11": {"agent": "اليزيد", "phone": "0666469265"},
    },
    "تاسريرت": {
        "1": {"agent": "كريم", "phone": "0673247646"},
        "2": {"agent": "مصطفى", "phone": "0662399349"},
        "3": {"agent": "مصطفى", "phone": "0662399349"},
        "4": {"agent": "بلحوس", "phone": "0639199353"},
        "5": {"agent": "بولودن", "phone": "0655513477"},
        "6": {"agent": "بولودن", "phone": "0655513477"},
        "7": {"agent": "وهمي", "phone": "0653801430"},
        "8": {"agent": "وهمي", "phone": "0653801430"},
        "9": {"agent": "وهمي", "phone": "0653801430"},
        "10": {"agent": "وهمي", "phone": "0653801430"},
        "11": {"agent": "بلحوس", "phone": "0639199353"},
    }
}

# -------------------------------------------------------------
# دالة تصدير ملف Excel بتنسيق عربي جميل
# -------------------------------------------------------------
def export_styled_excel(df, sheet_title="المعطيات"):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.sheet_view.rightToLeft = True

    cols_to_use = list(df.columns)
    ws.append(cols_to_use)

    for _, row in df.iterrows():
        ws.append([str(row[col]) if pd.notna(row[col]) else "-" for col in cols_to_use])

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(name="Calibri", size=14, bold=True, color="000000")
    data_font = Font(name="Calibri", size=13, bold=False, color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cols_to_use)):
        for cell in row:
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    wb.save(output)
    output.seek(0)
    return output.getvalue()

# -------------------------------------------------------------
# 1. قراءة واستخراج المكاتب والإحداثيات
# -------------------------------------------------------------
def load_offices_data(file_source):
    excel_data = pd.ExcelFile(file_source)
    sheets_map = {
        "أملن": "مكاتب التصويت أملن العمالة ",
        "تاسريرت": "مكاتب التصويت تسريرت العمالة",
        "تارسواط": "مكاتب التصويت ترسواط العمالة"
    }
    
    parsed_offices = []
    
    for commune_name, sheet_name in sheets_map.items():
        if sheet_name in excel_data.sheet_names:
            df = excel_data.parse(sheet_name)
            for _, row in df.iterrows():
                try:
                    c_name = str(row.iloc[2]).strip()
                    if c_name in ["أملن", "تاسريرت", "تارسواط"]:
                        num = str(int(row.iloc[4]))
                        center = str(row.iloc[5]).strip()
                        address = str(row.iloc[6]).strip()
                        
                        raw_lon = str(row.iloc[7]).replace(',', '.')
                        raw_lat = str(row.iloc[8]).replace(',', '.')
                        lon = float(raw_lon)
                        lat = float(raw_lat)
                        
                        parsed_offices.append({
                            "commune": c_name,
                            "num": num,
                            "center_name": center,
                            "name": f"مكتب {num} - {center}",
                            "address": address,
                            "lat": lat,
                            "lon": lon
                        })
                except (ValueError, TypeError, IndexError):
                    continue

    central_sheets = {
        "أملن": " رؤساء و أعضاء أملن المركزي ",
        "تاسريرت": " رؤساء و أعضاء تاسريرت المركزي1",
        "تارسواط": "رؤساء و أعضاء المركزي تارسواط1"
    }

    for commune_name, sheet_name in central_sheets.items():
        if sheet_name in excel_data.sheet_names:
            df_c = excel_data.parse(sheet_name)
            center_val = "المكتب المركزي"
            address_val = f"مقر جماعة {commune_name}"
            
            try:
                for idx in range(0, 10):
                    row_str = " ".join([str(val) for val in df_c.iloc[idx].values if pd.notna(val)])
                    if "مقر" in row_str or "جماعة" in row_str or "مركز" in row_str:
                        address_val = row_str
                        break
            except Exception:
                pass
                
            parsed_offices.append({
                "commune": commune_name,
                "num": "المركزي",
                "center_name": center_val,
                "name": f"المكتب المركزي - جماعة {commune_name}",
                "address": address_val,
                "lat": 29.7262,
                "lon": -8.9774
            })

    return pd.DataFrame(parsed_offices)

# -------------------------------------------------------------
# 2. قراءة واستخراج الأعضاء
# -------------------------------------------------------------
def load_members_data(file_source):
    excel_data = pd.ExcelFile(file_source)
    member_sheets = {
        "أملن": (" رؤساء و أعضاء المكاتب أملن2026", " رؤساء و أعضاء أملن المركزي "),
        "تاسريرت": ("رؤساء و أعضاء المكاتب تاسريرت1", " رؤساء و أعضاء تاسريرت المركزي1"),
        "تارسواط": ("رؤساء و أعضاء المكاتب تارسواط1", "رؤساء و أعضاء المركزي تارسواط1")
    }

    records = []

    def format_phone(phone_val):
        if pd.isna(phone_val) or str(phone_val).strip() in ["", "-"]:
            return "-"
        phone_str = str(phone_val).strip().replace(".0", "").replace(" ", "")
        if len(phone_str) == 9 and not phone_str.startswith("0"):
            return "0" + phone_str
        return phone_str

    for commune_name, (local_sheet, central_sheet) in member_sheets.items():
        if local_sheet in excel_data.sheet_names:
            df = excel_data.parse(local_sheet)
            for idx in range(8, len(df)):
                row = df.iloc[idx]
                try:
                    bureau_num = row.iloc[3]
                    if pd.notna(bureau_num) and str(bureau_num).strip().isdigit():
                        num = int(bureau_num)
                        roles = [
                            ("الرئيس", 4, 5, 6),
                            ("نائب الرئيس", 8, 9, 10),
                            ("العضو الأول", 12, 13, 14),
                            ("العضو الثاني", 16, 17, 18),
                            ("الكاتب", 20, 21, 22),
                            ("نائب العضو الأول", 24, 25, 26),
                            ("نائب العضو الثاني", 28, 29, 30),
                            ("نائب الكاتب", 32, 33, 34),
                        ]
                        for role_title, col_name, col_cin, col_phone in roles:
                            name = row.iloc[col_name] if len(row) > col_name else None
                            if pd.notna(name) and str(name).strip() != "":
                                records.append({
                                    "الجماعة": commune_name,
                                    "نوع المكتب": "مكتب تصويت محلي",
                                    "رقم المكتب": str(num),
                                    "الصفة": role_title,
                                    "الاسم الكامل": str(name).strip(),
                                    "CIN": str(row.iloc[col_cin]).strip() if pd.notna(row.iloc[col_cin]) else "-",
                                    "الهاتف": format_phone(row.iloc[col_phone]) if len(row) > col_phone else "-"
                                })
                except Exception:
                    continue

        if central_sheet in excel_data.sheet_names:
            df_c = excel_data.parse(central_sheet)
            for idx in range(8, len(df_c)):
                row = df_c.iloc[idx]
                try:
                    bureau_num = row.iloc[3]
                    if pd.notna(bureau_num):
                        roles = [
                            ("رئيس المكتب المركزي", 4, 5, 6),
                            ("نائب رئيس المكتب المركزي", 8, 9, 10),
                            ("العضو الأول", 12, 13, 14),
                            ("العضو الثاني", 16, 17, 18),
                            ("الكاتب", 20, 21, 22),
                            ("نائب العضو الأول", 24, 25, 26),
                            ("نائب العضو الثاني", 28, 29, 30),
                            ("نائب الكاتب", 32, 33, 34),
                        ]
                        for role_title, col_name, col_cin, col_phone in roles:
                            name = row.iloc[col_name] if len(row) > col_name else None
                            if pd.notna(name) and str(name).strip() != "":
                                records.append({
                                    "الجماعة": commune_name,
                                    "نوع المكتب": "المكتب المركزي",
                                    "رقم المكتب": "المركزي",
                                    "الصفة": role_title,
                                    "الاسم الكامل": str(name).strip(),
                                    "CIN": str(row.iloc[col_cin]).strip() if pd.notna(row.iloc[col_cin]) else "-",
                                    "الهاتف": format_phone(row.iloc[col_phone]) if len(row) > col_phone else "-"
                                })
                except Exception:
                    continue

    return pd.DataFrame(records)

# -------------------------------------------------------------
# 3. إدارة وحفظ اللوجستيك
# -------------------------------------------------------------
def load_logistics_data():
    if os.path.exists(LOGISTICS_FILE):
        try:
            return pd.read_excel(LOGISTICS_FILE)
        except Exception:
            pass
    return pd.DataFrame(columns=[
        "الجماعة", "رقم المكتب", "الربط الكهربائي", "عدد المعازل", 
        "الإضاءة الكافية", "عدد الكراسي", "عدد الطاولات", "عدد اللافتات", "ملاحظات"
    ])

def save_logistics_record(data_row):
    df_log = load_logistics_data()
    mask = (df_log["الجماعة"] == data_row["الجماعة"]) & (df_log["رقم المكتب"].astype(str) == str(data_row["رقم المكتب"]))
    if any(mask):
        df_log.loc[mask, list(data_row.keys())] = list(data_row.values())
    else:
        df_log = pd.concat([df_log, pd.DataFrame([data_row])], ignore_index=True)
    df_log.to_excel(LOGISTICS_FILE, index=False)

def delete_single_logistics_record(commune, office_num):
    df_log = load_logistics_data()
    df_log = df_log[~((df_log["الجماعة"] == commune) & (df_log["رقم المكتب"].astype(str) == str(office_num)))]
    df_log.to_excel(LOGISTICS_FILE, index=False)

# -------------------------------------------------------------
# 4. إدارة وحفظ بيانات التنقل والسيارات (مع التعبئة التلقائية لأملن وتاسريرت)
# -------------------------------------------------------------
def load_transport_data(df_offices_ref):
    if os.path.exists(TRANSPORT_FILE):
        try:
            df_trans = pd.read_excel(TRANSPORT_FILE)
            return df_trans.fillna("").astype(str)
        except Exception:
            pass

    data = []
    if not df_offices_ref.empty:
        for _, row in df_offices_ref.iterrows():
            comm = str(row.get("commune", ""))
            num = str(row.get("num", ""))
            
            # الجلب التلقائي للأعوان المعرفين لجماعتي أملن وتاسريرت
            agent_info = COMMUNES_AGENTS.get(comm, {}).get(num, {"agent": "", "phone": ""})
            
            data.append({
                "الجماعة": comm,
                "صاحب السيارة": agent_info["agent"],
                "نوع السيارة": "",
                "برمجتها للمكتب رقم": num,
                "العنوان / المسار": str(row.get("address", "")),
                "رقم الهاتف": agent_info["phone"],
                "ملاحظات التنقل": ""
            })
    return pd.DataFrame(data).fillna("").astype(str)

def save_transport_data(df_trans):
    df_trans.fillna("").astype(str).to_excel(TRANSPORT_FILE, index=False)

# -------------------------------------------------------------
# إعداد Session State
# -------------------------------------------------------------
if "df_members" not in st.session_state:
    if os.path.exists(EXCEL_FILE):
        st.session_state.df_offices = load_offices_data(EXCEL_FILE)
        st.session_state.df_members = load_members_data(EXCEL_FILE)
    else:
        st.session_state.df_offices = pd.DataFrame()
        st.session_state.df_members = pd.DataFrame()

if "df_transport" not in st.session_state:
    st.session_state.df_transport = load_transport_data(st.session_state.df_offices)

df_offices = st.session_state.df_offices
df_members = st.session_state.df_members

# -------------------------------------------------------------
# القائمة الجانبية
# -------------------------------------------------------------
st.sidebar.header("⚙️ التصفية والتحكم")
commune_list = ["الكل", "أملن", "تاسريرت", "تارسواط"]
selected_commune = st.sidebar.selectbox("اختر الجماعة الترابية:", commune_list)

filtered_offices = df_offices if selected_commune == "الكل" else df_offices[df_offices["commune"] == selected_commune]
filtered_members = df_members if selected_commune == "الكل" else df_members[df_members["الجماعة"] == selected_commune]

# -------------------------------------------------------------
# تبويبات التطبيق
# -------------------------------------------------------------
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "👥 رؤساء المكاتب ونوابهم والأعضاء ونوابهم", 
    "🚗 التنقل والوسائل",
    "🗺️ الخريطة والمواقع", 
    "🏢 التجهيزات واللوجستيك",
    "📋 البطاقة التقنية"
])

# -------------------------------------------------------------
# TAB 1: الأعضاء
# -------------------------------------------------------------
with tab1:
    st.header(f"👥 لائحة رؤساء المكاتب ونوابهم والأعضاء ونوابهم ({selected_commune})")
    
    if not filtered_members.empty:
        unique_offices = filtered_members["رقم المكتب"].unique()
        numeric_offices = sorted([int(x) for x in unique_offices if str(x).isdigit()])
        has_central = any(str(x) == "المركزي" for x in unique_offices)
        
        offices_in_members = ["الكل"] + [str(x) for x in numeric_offices]
        if has_central:
            offices_in_members.append("المركزي")
            
        selected_office_num = st.selectbox("تصفية حسب رقم المكتب:", offices_in_members, key="tab1_office")
        
        display_df = filtered_members.copy()
        if selected_office_num != "الكل":
            display_df = display_df[display_df["رقم المكتب"].astype(str) == str(selected_office_num)]

        st.info("💡 **يمكنك تعديل أي خلية مباشرة في الجدول أدناه، إضافة صفوف جديدة، أو مسح صفوف.**")
        
        edited_df = st.data_editor(
            display_df, 
            use_container_width=True, 
            num_rows="dynamic", 
            key=f"editor_{selected_commune}_{selected_office_num}"
        )
        
        col_save, col_empty = st.columns([1, 2])
        with col_save:
            if st.button("💾 تثبيت وحفظ التعديلات نهائياً", use_container_width=True, type="primary", key="btn_save_members"):
                if selected_office_num == "الكل" and selected_commune == "الكل":
                    st.session_state.df_members = edited_df
                else:
                    st.session_state.df_members.update(edited_df)
                
                st.success("✅ تم تثبيت التعديلات بنجاح!")
                st.rerun()

        st.markdown("---")
        st.subheader("📥 تصدير وتحميل القائمة المعدلة")
        
        col_excel, col_print = st.columns(2)
        
        with col_excel:
            excel_bytes = export_styled_excel(edited_df, "الأعضاء")
            st.download_button(
                label="📊 تحميل القائمة المعدلة (Excel)",
                data=excel_bytes,
                file_name=f"لائحة_الأعضاء_{selected_commune}_{selected_office_num}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col_print:
            html_table = edited_df.to_html(index=False, classes='table', justify='center')
            html_document = f"""
            <!DOCTYPE html>
            <html dir="rtl" lang="ar">
            <head>
                <meta charset="utf-8">
                <style>
                    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; text-align: center; padding: 20px; }}
                    h2 {{ color: #1f77b4; margin-bottom: 5px; }}
                    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                    th, td {{ border: 1px solid #000; padding: 10px; text-align: center; font-size: 13px; }}
                    th {{ background-color: #FFFF00; color: #000; font-size: 15px; }}
                    @media print {{ button {{ display: none; }} }}
                </style>
            </head>
            <body>
                <h2>لائحة الأعضاء - جماعة {selected_commune} (مكتب {selected_office_num})</h2>
                {html_table}
                <br>
                <button onclick="window.print()" style="padding: 10px 20px; background-color: #28a745; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 16px;">🖨️ طباعة / حفظ كـ PDF</button>
            </body>
            </html>
            """
            
            st.download_button(
                label="🖨️ طباعة / تحميل PDF",
                data=html_document,
                file_name=f"طباعة_الأعضاء_{selected_commune}_{selected_office_num}.html",
                mime="text/html",
                use_container_width=True
            )
    else:
        st.warning("لا توجد بيانات متوفرة.")

# -------------------------------------------------------------
# TAB 2: التنقل والوسائل
# -------------------------------------------------------------
with tab2:
    st.header(f"🚗 برمجيات وسائط النقل والتنقل ({selected_commune})")
    st.caption("إدارة حافلات وسيارات النقل: صاحب السيارة - نوعها - برمجتها للمكتب - العنوان والمسار")

    df_trans_display = st.session_state.df_transport.copy()
    
    if selected_commune != "الكل" and "الجماعة" in df_trans_display.columns:
        df_trans_display = df_trans_display[df_trans_display["الجماعة"] == selected_commune]

    df_trans_display = df_trans_display.fillna("").astype(str)

    st.info("💡 **يمكنك تعديل أي معلومة مباشرة في الجدول أو إضافة مركبات جديدة.**")

    edited_trans_df = st.data_editor(
        df_trans_display,
        use_container_width=True,
        num_rows="dynamic",
        key=f"editor_trans_{selected_commune}",
        column_config={
            "الجماعة": st.column_config.SelectboxColumn("الجماعة", options=["أملن", "تاسريرت", "تارسواط"]),
            "صاحب السيارة": st.column_config.TextColumn("صاحب السيارة"),
            "نوع السيارة": st.column_config.TextColumn("نوع السيارة"),
            "برمجتها للمكتب رقم": st.column_config.TextColumn("برمجتها للمكتب رقم"),
            "العنوان / المسار": st.column_config.TextColumn("العنوان / المسار"),
            "رقم الهاتف": st.column_config.TextColumn("رقم الهاتف"),
            "ملاحظات التنقل": st.column_config.TextColumn("ملاحظات التنقل")
        }
    )

    col_trans_save, col_trans_export = st.columns(2)
    
    with col_trans_save:
        if st.button("💾 حفظ برمجية التنقل نهائياً", use_container_width=True, type="primary", key="btn_save_trans"):
            if selected_commune == "الكل":
                st.session_state.df_transport = edited_trans_df.fillna("").astype(str)
            else:
                other_data = st.session_state.df_transport[st.session_state.df_transport["الجماعة"] != selected_commune]
                st.session_state.df_transport = pd.concat([other_data, edited_trans_df], ignore_index=True).fillna("").astype(str)
            
            save_transport_data(st.session_state.df_transport)
            st.success("✅ تم حفظ بيانات وسائل النقل والتنقل بنجاح!")
            st.rerun()

    with col_trans_export:
        excel_trans_bytes = export_styled_excel(edited_trans_df, "برمجية التنقل")
        st.download_button(
            label="📊 تحميل برمجية التنقل (Excel)",
            data=excel_trans_bytes,
            file_name=f"برمجية_التنقل_{selected_commune}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# -------------------------------------------------------------
# TAB 3: الخريطة (Satellite View + Auto Boundaries)
# -------------------------------------------------------------
with tab3:
    st.header(f"🗺️ خريطة مواقع مكاتب التصويت ({selected_commune})")
    
    if not filtered_offices.empty:
        map_df = filtered_offices.copy()
        
        map_df['lat'] = map_df['lat'].astype(str).str.replace(',', '.')
        map_df['lon'] = map_df['lon'].astype(str).str.replace(',', '.')
        
        map_df['lat'] = pd.to_numeric(map_df['lat'], errors='coerce')
        map_df['lon'] = pd.to_numeric(map_df['lon'], errors='coerce')
        
        valid_offices = map_df.dropna(subset=["lat", "lon"])
        
        if not valid_offices.empty:
            avg_lat = valid_offices["lat"].mean()
            avg_lon = valid_offices["lon"].mean()
            
            m = folium.Map(
                location=[avg_lat, avg_lon], 
                zoom_start=11, 
                max_zoom=20
            )
            
            folium.TileLayer(
                tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                attr='Esri',
                name='صور القمر الصناعي (Satellite)',
                max_zoom=20,
                overlay=False,
                control=True
            ).add_to(m)

            folium.TileLayer(
                tiles='OpenStreetMap',
                name='الخريطة العادية (Street Map)',
                max_zoom=19,
                overlay=False,
                control=True
            ).add_to(m)

            folium.LayerControl(position='topright').add_to(m)
            
            bounds = []
            marker_cluster = MarkerCluster().add_to(m)

            for _, row in valid_offices.iterrows():
                lat, lon = float(row["lat"]), float(row["lon"])
                bounds.append([lat, lon])
                
                is_central = (str(row["num"]) == "المركزي")
                marker_color = "red" if is_central else "blue"
                icon_type = "star" if is_central else "info-sign"
                
                popup_html = f"""
                <div dir="rtl" style="font-family: Arial; min-width: 180px; text-align: right;">
                    <h4 style="color: #1f77b4; margin-bottom: 5px;">جماعة {row['commune']}</h4>
                    <b>مكتب رقم: {row['num']}</b><br>
                    <b>الاسم:</b> {row['center_name'] if 'center_name' in row else row['name']}<br>
                    <small><b>العنوان:</b> {row['address']}</small>
                </div>
                """
                
                folium.Marker(
                    [lat, lon],
                    popup=folium.Popup(popup_html, max_width=300),
                    tooltip=f"مكتب {row['num']} - جماعة {row['commune']}",
                    icon=folium.Icon(color=marker_color, icon=icon_type)
                ).add_to(marker_cluster)
            
            if len(bounds) > 1:
                m.fit_bounds(bounds, padding=[30, 30])
            
            st_folium(m, width="100%", height=600, returned_objects=[])
            
            invalid_count = len(map_df) - len(valid_offices)
            if invalid_count > 0:
                st.warning(f"⚠️ يوجد {invalid_count} مكتب/مكاتب لم تظهر على الخريطة بسبب إحداثيات غير صحيحة.")
        else:
            st.error("❌ لا توجد إحداثيات صالحة لعرض الخريطة.")
    else:
        st.warning("لا توجد بيانات مواقع لعرضها على الخريطة.")

# -------------------------------------------------------------
# TAB 4: التجهيزات واللوجستيك (مع زر التعديل)
# -------------------------------------------------------------
with tab4:
    st.header("🏢 جرد وإدارة التجهيزات واللوجستيك")
    
    col_sel_c, col_sel_o = st.columns(2)
    with col_sel_c:
        logistics_commune = st.selectbox("اختر الجماعة:", ["أملن", "تاسريرت", "تارسواط"], key="log_c")

    available_offices = []
    if not df_offices.empty:
        commune_offices = df_offices[df_offices["commune"] == logistics_commune]
        if not commune_offices.empty:
            local_nums = sorted([int(x) for x in commune_offices["num"].unique() if str(x).isdigit()])
            available_offices = [str(x) for x in local_nums]
            if "المركزي" in commune_offices["num"].values:
                available_offices.append("المركزي")

    if not available_offices and not df_members.empty:
        commune_m = df_members[df_members["الجماعة"] == logistics_commune]
        if not commune_m.empty:
            local_nums = sorted([int(x) for x in commune_m["رقم المكتب"].unique() if str(x).isdigit()])
            available_offices = [str(x) for x in local_nums]
            if "المركزي" in commune_m["رقم المكتب"].values:
                available_offices.append("المركزي")

    if not available_offices:
        available_offices = ["1"]

    with col_sel_o:
        logistics_office_num = st.selectbox(
            "اختر رقم المكتب:",
            options=available_offices,
            key=f"log_o_select_{logistics_commune}"
        )

    df_logistics_existing = load_logistics_data()
    existing_row = df_logistics_existing[
        (df_logistics_existing["الجماعة"] == logistics_commune) & 
        (df_logistics_existing["رقم المكتب"].astype(str) == str(logistics_office_num))
    ]
    
    is_editing = not existing_row.empty
    
    if is_editing:
        st.info(f"ℹ️ توجد بيانات لوجستية مسجلة مسبقاً لمكتب ({logistics_office_num} - {logistics_commune}). يمكنك تعديلها ثم الضغط على زر التعديل والتحديث.")
        v_elec = bool(existing_row["الربط الكهربائي"].values[0] == "نعم")
        v_light = bool(existing_row["الإضاءة الكافية"].values[0] == "نعم")
        
        try:
            v_barriers = int(float(existing_row["عدد المعازل"].values[0]))
            if v_barriers not in [1, 2]: v_barriers = 1
        except Exception:
            v_barriers = 1
            
        v_chairs = int(float(existing_row["عدد الكراسي"].values[0])) if pd.notna(existing_row["عدد الكراسي"].values[0]) else 10
        v_tables = int(float(existing_row["عدد الطاولات"].values[0])) if pd.notna(existing_row["عدد الطاولات"].values[0]) else 4
        v_signs = int(float(existing_row["عدد اللافتات"].values[0])) if pd.notna(existing_row["عدد اللافتات"].values[0]) else 2
        v_notes = str(existing_row["ملاحظات"].values[0]) if pd.notna(existing_row["ملاحظات"].values[0]) else ""
    else:
        v_elec = True
        v_light = True
        v_barriers = 1
        v_chairs = 10
        v_tables = 4
        v_signs = 2
        v_notes = ""

    col1, col2 = st.columns(2)
    with col1:
        elec = st.checkbox("توفر الربط الكهربائي", value=v_elec, key=f"c_elec_{logistics_commune}_{logistics_office_num}")
        lighting = st.checkbox("توفر الإضاءة الكافية", value=v_light, key=f"c_light_{logistics_commune}_{logistics_office_num}")
        barriers_count = st.number_input("عدد المعازل (1 أو 2 فقط):", min_value=1, max_value=2, value=v_barriers, step=1, key=f"c_bar_{logistics_commune}_{logistics_office_num}")
        
    with col2:
        chairs = st.number_input("عدد الكراسي", min_value=0, value=v_chairs, key=f"c_ch_{logistics_commune}_{logistics_office_num}")
        tables = st.number_input("عدد الطاولات", min_value=0, value=v_tables, key=f"c_tb_{logistics_commune}_{logistics_office_num}")
        signs = st.number_input("عدد اللافتات", min_value=0, value=v_signs, key=f"c_sg_{logistics_commune}_{logistics_office_num}")

    notes = st.text_area("ملاحظات:", value=v_notes if v_notes != "nan" else "", key=f"c_nt_{logistics_commune}_{logistics_office_num}")

    st.markdown("---")
    
    btn_col1, btn_col2 = st.columns(2)
    
    data_to_save = {
        "الجماعة": logistics_commune,
        "رقم المكتب": logistics_office_num,
        "الربط الكهربائي": "نعم" if elec else "لا",
        "عدد المعازل": barriers_count,
        "الإضاءة الكافية": "نعم" if lighting else "لا",
        "عدد الكراسي": chairs,
        "عدد الطاولات": tables,
        "عدد اللافتات": signs,
        "ملاحظات": notes
    }

    if is_editing:
        with btn_col1:
            if st.button("✏️ تعديل وتحديث بيانات هذا المكتب", use_container_width=True, type="primary", key="btn_edit_log"):
                save_logistics_record(data_to_save)
                st.success(f"✅ تم تعديل وتحديث تجهيزات مكتب {logistics_office_num} - جماعة {logistics_commune} بنجاح!")
                st.rerun()
        with btn_col2:
            if st.button("🗑️ مسح بيانات هذا المكتب", use_container_width=True, key="btn_del_log_curr"):
                delete_single_logistics_record(logistics_commune, logistics_office_num)
                st.warning("⚠️ تم مسح سجل التجهيزات لهذا المكتب.")
                st.rerun()
    else:
        with btn_col1:
            if st.button("💾 حفظ التجهيزات لهذا المكتب", use_container_width=True, type="primary", key="btn_save_log"):
                save_logistics_record(data_to_save)
                st.success(f"✅ تم حفظ تجهيزات مكتب {logistics_office_num} - جماعة {logistics_commune} بنجاح!")
                st.rerun()

    st.markdown("---")
    st.subheader(f"📋 السجل الشامل للجرد اللوجستي - جماعة {logistics_commune}")
    
    df_logistics_all = load_logistics_data()
    
    if not df_logistics_all.empty:
        df_logistics_filtered = df_logistics_all[df_logistics_all["الجماعة"] == logistics_commune]
    else:
        df_logistics_filtered = pd.DataFrame()

    if not df_logistics_filtered.empty:
        for idx, row in df_logistics_filtered.iterrows():
            c_name = row["الجماعة"]
            o_num = row["رقم المكتب"]
            
            barriers_raw = row.get("عدد المعازل", "1")
            barriers_val = "1" if pd.isna(barriers_raw) else str(barriers_raw).replace('.0', '').strip()

            chairs_val = str(row.get('عدد الكراسي', '-')).replace('.0', '')
            tables_val = str(row.get('عدد الطاولات', '-')).replace('.0', '')
            signs_val = str(row.get('عدد اللافتات', '-')).replace('.0', '')
            
            with st.expander(f"📍 **جماعة {c_name} - مكتب رقم {o_num}**", expanded=True):
                col_data, col_action = st.columns([5, 1])
                
                with col_data:
                    st.write(f"⚡ كهرباء: {row.get('الربط الكهربائي', '-')} | 🚪 عدد المعازل: {barriers_val} | 💡 إضاءة: {row.get('الإضاءة الكافية', '-')}")
                    st.write(f"🪑 كراسي: {chairs_val} | طاولات: {tables_val} | لافتات: {signs_val}")
                    if pd.notna(row.get("ملاحظات")) and str(row.get("ملاحظات")).strip() not in ["", "nan"]:
                        st.caption(f"📝 ملاحظات: {row['ملاحظات']}")
                
                with col_action:
                    if st.button("🗑️ مسح", key=f"del_{c_name}_{o_num}_{idx}", use_container_width=True):
                        delete_single_logistics_record(c_name, o_num)
                        st.success("تم المسح بنجاح!")
                        st.rerun()
    else:
        st.info(f"لا توجد بيانات لوجستية مسجلة لجماعة {logistics_commune}.")

# -------------------------------------------------------------
# TAB 5: البطاقة التقنية
# -------------------------------------------------------------
with tab5:
    st.markdown("<h2 style='text-align: center; color: #1E3A8A;'>دائرة تافراوت - قيادة أملن</h2>", unsafe_allow_html=True)
    st.markdown("<h3 style='text-align: center; color: #1E3A8A; margin-bottom: 25px;'>بطاقة تقنية حول مكتب التصويت</h3>", unsafe_allow_html=True)

    col_sel_c, col_sel_o, col_label = st.columns([2, 2, 2])
    
    with col_sel_c:
        card_commune = st.selectbox("اختيار الجماعة:", ["أملن", "تاسريرت", "تارسواط"], key="card_c_select")

    card_office_options = []
    if not df_offices.empty:
        c_offices = df_offices[df_offices["commune"] == card_commune]
        if not c_offices.empty:
            num_list = sorted([int(x) for x in c_offices["num"].unique() if str(x).isdigit()])
            card_office_options = [str(x) for x in num_list]
            if "المركزي" in c_offices["num"].values:
                card_office_options.append("المركزي")

    if not card_office_options and not df_members.empty:
        c_mem = df_members[df_members["الجماعة"] == card_commune]
        if not c_mem.empty:
            num_list = sorted([int(x) for x in c_mem["رقم المكتب"].unique() if str(x).isdigit()])
            card_office_options = [str(x) for x in num_list]
            if "المركزي" in c_mem["رقم المكتب"].values:
                card_office_options.append("المركزي")

    if not card_office_options:
        card_office_options = ["1"]

    with col_sel_o:
        card_office_num = st.selectbox("اختيار رقم المكتب:", card_office_options, key="card_o_select")

    with col_label:
        st.markdown(f"<div style='text-align: left; padding-top: 28px; color: #888; font-style: italic;'>(تتغير البيانات تلقائياً) &nbsp;&nbsp;&nbsp; <b>{card_commune} - مكتب {card_office_num}</b></div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    info_row = df_offices[(df_offices["commune"] == card_commune) & (df_offices["num"].astype(str) == str(card_office_num))]
    address_val = info_row.iloc[0]["address"] if not info_row.empty and "address" in info_row.columns else "-"
    center_val = info_row.iloc[0]["center_name"] if not info_row.empty and "center_name" in info_row.columns else "-"

    mem_df = df_members[(df_members["الجماعة"] == card_commune) & (df_members["رقم المكتب"].astype(str) == str(card_office_num))]
    
    president_name = "-"
    m1_name, m2_name, m3_name = "-", "-", "-"
    
    if not mem_df.empty:
        pres_row = mem_df[mem_df["الصفة"].str.contains("رئيس", na=False) & ~mem_df["الصفة"].str.contains("نائب", na=False)]
        if not pres_row.empty:
            president_name = pres_row.iloc[0]["الاسم الكامل"]

        m1_row = mem_df[mem_df["الصفة"].str.contains("الأول", na=False) & ~mem_df["الصفة"].str.contains("نائب", na=False)]
        if not m1_row.empty:
            m1_name = m1_row.iloc[0]["الاسم الكامل"]

        m2_row = mem_df[mem_df["الصفة"].str.contains("الثاني", na=False) & ~mem_df["الصفة"].str.contains("نائب", na=False)]
        if not m2_row.empty:
            m2_name = m2_row.iloc[0]["الاسم الكامل"]

        m3_row = mem_df[mem_df["الصفة"].str.contains("الكاتب", na=False) & ~mem_df["الصفة"].str.contains("نائب", na=False)]
        if not m3_row.empty:
            m3_name = m3_row.iloc[0]["الاسم الكامل"]

    df_log_all = load_logistics_data()
    log_row = df_log_all[(df_log_all["الجماعة"].astype(str) == str(card_commune)) & (df_log_all["رقم المكتب"].astype(str) == str(card_office_num))] if not df_log_all.empty else pd.DataFrame()

    elec_val = log_row.iloc[0].get("الربط الكهربائي", "0") if not log_row.empty else "0"
    chairs_val = str(log_row.iloc[0].get("عدد الكراسي", "0")).replace('.0', '') if not log_row.empty else "0"
    tables_val = str(log_row.iloc[0].get("عدد الطاولات", "0")).replace('.0', '') if not log_row.empty else "0"
    barriers_val = str(log_row.iloc[0].get("عدد المعازل", "0")).replace('.0', '') if not log_row.empty else "0"
    
    # استخراج بيانات العون المكلف (الأولوية لقاموس الجماعات المباشر ثم لجدول التنقل)
    agent_name = "-"
    agent_phone = "-"
    agent_date = "-"

    if card_commune in COMMUNES_AGENTS and str(card_office_num) in COMMUNES_AGENTS[card_commune]:
        agent_name = COMMUNES_AGENTS[card_commune][str(card_office_num)]["agent"]
        agent_phone = COMMUNES_AGENTS[card_commune][str(card_office_num)]["phone"]
    else:
        df_trans_all = st.session_state.df_transport
        trans_row = df_trans_all[(df_trans_all["الجماعة"].astype(str) == str(card_commune)) & (df_trans_all["برمجتها للمكتب رقم"].astype(str) == str(card_office_num))] if not df_trans_all.empty else pd.DataFrame()
        if not trans_row.empty:
            agent_name = trans_row.iloc[0].get("صاحب السيارة", "-")
            agent_phone = trans_row.iloc[0].get("رقم الهاتف", "-")
            agent_date = trans_row.iloc[0].get("ملاحظات التنقل", "-")

    card_style = """
    <style>
        .tech-card-table {
            width: 100%;
            border-collapse: collapse;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            direction: rtl;
            margin-bottom: 20px;
        }
        .tech-card-table th.section-header {
            background-color: #1A4D7C;
            color: white;
            text-align: right;
            padding: 10px 15px;
            font-size: 16px;
            font-weight: bold;
            border: 1px solid #1A4D7C;
        }
        .tech-card-table td {
            border: 1px solid #A6B9D0;
            padding: 8px 12px;
            text-align: center;
            font-size: 14px;
        }
        .tech-card-table td.label-cell {
            background-color: #E2EBF5;
            font-weight: bold;
            color: #000;
            width: 18%;
        }
        .tech-card-table td.value-cell {
            background-color: #FFFFFF;
            color: #000;
        }
    </style>
    """

    card_table_body = f"""
    <table class="tech-card-table">
        <tr>
            <th colspan="6" class="section-header">البيانات العامة</th>
        </tr>
        <tr>
            <td class="label-cell">الجماعة:</td>
            <td class="value-cell" style="width: 25%;">{card_commune}</td>
            <td class="label-cell">رقم الدائرة الانتخابية:</td>
            <td class="value-cell" style="width: 15%;">{card_office_num}</td>
            <td class="label-cell">رقم مكتب التصويت:</td>
            <td class="value-cell" style="width: 15%;">{card_office_num}</td>
        </tr>
        <tr>
            <td class="label-cell">عنوان مكتب التصويت:</td>
            <td colspan="5" class="value-cell" style="text-align: right; padding-right: 15px;">{address_val if address_val != "-" else center_val}</td>
        </tr>
        <tr>
            <th colspan="6" class="section-header">أعضاء مكتب التصويت</th>
        </tr>
        <tr>
            <td class="label-cell">رئيس مكتب التصويت:</td>
            <td class="value-cell" style="width: 30%;">{president_name}</td>
            <td class="label-cell">عضو المكتب 1:</td>
            <td colspan="3" class="value-cell">{m1_name}</td>
        </tr>
        <tr>
            <td colspan="2" rowspan="2" style="background-color: #F9FAFB;"></td>
            <td class="label-cell">عضو المكتب 2:</td>
            <td colspan="3" class="value-cell">{m2_name}</td>
        </tr>
        <tr>
            <td class="label-cell">عضو المكتب 3:</td>
            <td colspan="3" class="value-cell">{m3_name}</td>
        </tr>
        <tr>
            <th colspan="6" class="section-header">الوسائل اللوجستيكية والحالة العامة</th>
        </tr>
        <tr>
            <td class="label-cell">الكهرباء:</td>
            <td class="value-cell">{elec_val}</td>
            <td class="label-cell">الكراسي:</td>
            <td class="value-cell">{chairs_val}</td>
            <td class="label-cell">الطاولات:</td>
            <td class="value-cell">{tables_val}</td>
        </tr>
        <tr>
            <td class="label-cell">النظافة:</td>
            <td class="value-cell">0</td>
            <td class="label-cell">التهوية:</td>
            <td class="value-cell">0</td>
            <td class="label-cell">الولوجية:</td>
            <td class="value-cell">0</td>
        </tr>
        <tr>
            <td class="label-cell">الصباغة / الحالة:</td>
            <td class="value-cell">0</td>
            <td class="label-cell">عدد المعازل:</td>
            <td class="value-cell">{barriers_val}</td>
            <td colspan="2" style="background-color: #F9FAFB;"></td>
        </tr>
        <tr>
            <th colspan="6" class="section-header">العون المكلف والتواصل</th>
        </tr>
        <tr>
            <td class="label-cell">العون المكلف:</td>
            <td class="value-cell">{agent_name}</td>
            <td class="label-cell">رقم الهاتف:</td>
            <td class="value-cell">{agent_phone}</td>
            <td class="label-cell">تاريخ التواصل:</td>
            <td class="value-cell">{agent_date}</td>
        </tr>
    </table>
    """

    st.markdown(card_style + card_table_body, unsafe_allow_html=True)

    print_card_html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
    <meta charset="utf-8">
    <title>بطاقة تقنية - مكتب {card_office_num}</title>
    {card_style}
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; padding: 30px; text-align: center; }}
        h2, h3 {{ color: #1A4D7C; margin: 5px; }}
        @media print {{ button {{ display: none; }} }}
    </style>
</head>
<body>
    <h2>دائرة تافراوت - قيادة أملن</h2>
    <h3>بطاقة تقنية حول مكتب التصويت رقم {card_office_num} - جماعة {card_commune}</h3>
    <br>
    {card_table_body}
    <br><br>
    <button onclick="window.print()" style="padding: 10px 25px; background-color: #1A4D7C; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; font-weight: bold;">🖨️ طباعة البطاقة التقنية</button>
</body>
</html>"""
    
    st.download_button(
        label="🖨️ طباعة البطاقة التقنية لهذا المكتب (PDF/HTML)",
        data=print_card_html,
        file_name=f"بطاقة_تقنية_مكتب_{card_office_num}_{card_commune}.html",
        mime="text/html",
        use_container_width=True
    )
